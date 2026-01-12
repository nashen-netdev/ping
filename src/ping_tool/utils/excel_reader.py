"""
Excel 文件读取模块，支持读取单元格格式（颜色、删除线等）
"""
import pandas as pd
import openpyxl
from typing import Optional, Dict, List, Tuple
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill


def read_excel_data_only(file_path: str, sheet_name: str) -> pd.DataFrame:
    """
    读取 Excel 文件数据
    
    Args:
        file_path: Excel 文件路径
        sheet_name: sheet 页名称
        
    Returns:
        pd.DataFrame: 数据框
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        return df
    except Exception as e:
        print(f"读取 Excel 文件失败: {e}")
        raise


def get_cell_style_info(file_path: str, sheet_name: str, row_idx: int, col_idx: int) -> Dict:
    """
    获取单元格的样式信息（颜色、删除线等）
    注意：此函数可能因为 Excel 文件格式问题而失败
    
    Args:
        file_path: Excel 文件路径
        sheet_name: sheet 页名称
        row_idx: 行索引（从1开始）
        col_idx: 列索引（从1开始）
        
    Returns:
        dict: 包含样式信息的字典
            {
                'fill_color': 'RRGGBB' or None,
                'font_strikethrough': bool,
                'font_color': 'RRGGBB' or None
            }
    """
    try:
        # 尝试只读取样式信息
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False)
        ws = wb[sheet_name]
        cell = ws.cell(row=row_idx, column=col_idx)
        
        style_info = {
            'fill_color': None,
            'font_strikethrough': False,
            'font_color': None
        }
        
        # 获取填充颜色
        if cell.fill and cell.fill.patternType:
            if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                    # 去掉透明度通道（前两位）
                    style_info['fill_color'] = str(cell.fill.fgColor.rgb)[-6:]
        
        # 获取字体样式
        if cell.font:
            style_info['font_strikethrough'] = bool(cell.font.strike)
            if hasattr(cell.font, 'color') and cell.font.color:
                if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                    style_info['font_color'] = str(cell.font.color.rgb)[-6:]
        
        wb.close()
        return style_info
    except Exception as e:
        # 如果读取样式失败，返回默认值
        return {
            'fill_color': None,
            'font_strikethrough': False,
            'font_color': None
        }


def get_sheet_style_map(file_path: str, sheet_name: str, max_rows: int = 1000) -> Dict[Tuple[int, int], Dict]:
    """
    批量获取整个 sheet 的样式信息（优化性能）
    
    Args:
        file_path: Excel 文件路径
        sheet_name: sheet 页名称
        max_rows: 最大读取行数
        
    Returns:
        dict: {(row, col): style_info}
    """
    style_map = {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False)
        ws = wb[sheet_name]
        
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows), 1):
            for col_idx, cell in enumerate(row, 1):
                style_info = {
                    'fill_color': None,
                    'font_strikethrough': False,
                    'font_color': None
                }
                
                # 获取填充颜色
                if cell.fill and cell.fill.patternType:
                    if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                        if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                            style_info['fill_color'] = str(cell.fill.fgColor.rgb)[-6:]
                
                # 获取字体样式
                if cell.font:
                    style_info['font_strikethrough'] = bool(cell.font.strike)
                    if hasattr(cell.font, 'color') and cell.font.color:
                        if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                            style_info['font_color'] = str(cell.font.color.rgb)[-6:]
                
                # 只存储有样式的单元格
                if any(style_info.values()):
                    style_map[(row_idx, col_idx)] = style_info
        
        wb.close()
        print(f"成功读取 {len(style_map)} 个有样式的单元格")
    except Exception as e:
        print(f"警告: 无法读取单元格样式信息: {e}")
        print("将跳过颜色和删除线过滤功能")
    
    return style_map


def is_green_cell(style_info: Dict) -> bool:
    """
    判断单元格是否为绿色
    
    Args:
        style_info: 样式信息字典
        
    Returns:
        bool: 是否为绿色单元格
    """
    if not style_info or not style_info.get('fill_color'):
        return False
    
    color = style_info['fill_color'].upper()
    # 常见的绿色值（可以根据实际情况调整）
    green_colors = [
        'C6EFCE',  # 浅绿色
        '00B050',  # Excel 标准绿色
        '92D050',  # 亮绿色
        'E2EFDA',  # 很浅的绿色
        '70AD47',  # 深绿色
    ]
    
    # 检查是否匹配已知的绿色
    for green in green_colors:
        if color == green:
            return True
    
    # 通用绿色检测：R 分量小，G 分量大
    try:
        r = int(color[0:2], 16)
        g = int(color[2:4], 16)
        b = int(color[4:6], 16)
        # 绿色的特征：G > R 且 G > B
        if g > r and g > b and g > 100:
            return True
    except:
        pass
    
    return False


def read_network_security_ips(file_path: str,
                               sheet_name: str = 'network&security',
                               ip_column: str = 'MGMT',
                               filter_color: Optional[str] = None,
                               exclude_strikethrough: bool = True) -> List[Dict]:
    """
    从指定 sheet 读取网络和安全设备的 IP 地址和 hostname
    
    Args:
        file_path: Excel 文件路径
        sheet_name: Sheet 页名称
        ip_column: 要 ping 的列名（'MGMT' 或 'IPMI'）
        filter_color: 过滤颜色，'green' 表示只读取绿色单元格，None 表示不过滤
        exclude_strikethrough: 是否排除删除线单元格
        
    Returns:
        list: [{'ip': '...', 'hostname': '...', 'row': ...}, ...]
    """
    
    # 使用多种方式尝试读取数据
    print(f"正在读取 {file_path} 的 {sheet_name} sheet...")
    df = read_excel_data_only(file_path, sheet_name)
    
    # 查找表头行（包含指定的 IP 列和 hostname）
    header_row_idx = None
    ip_col_idx = None
    hostname_col_idx = None
    
    for idx, row in df.iterrows():
        row_values = [str(x).strip() if pd.notna(x) else '' for x in row.values]
        if ip_column in row_values and 'hostname' in row_values:
            header_row_idx = idx
            ip_col_idx = row_values.index(ip_column)
            hostname_col_idx = row_values.index('hostname')
            print(f"找到表头行: 第{header_row_idx}行")
            print(f"{ip_column} 列索引: {ip_col_idx}, hostname 列索引: {hostname_col_idx}")
            break
    
    if header_row_idx is None:
        raise ValueError(f"在 {sheet_name} sheet 中未找到包含 '{ip_column}' 和 'hostname' 的表头行")
    
    # 尝试读取样式信息（如果需要过滤颜色或删除线）
    style_map = {}
    if filter_color or exclude_strikethrough:
        print("正在读取单元格样式信息...")
        style_map = get_sheet_style_map(file_path, sheet_name, max_rows=len(df))
    
    # 读取数据行
    results = []
    data_start_row = header_row_idx + 1
    
    print(f"开始从第{data_start_row}行读取数据...")
    for idx in range(data_start_row, len(df)):
        row = df.iloc[idx]
        ip_address = row.iloc[ip_col_idx]
        hostname = row.iloc[hostname_col_idx]
        
        # 跳过空 IP
        if pd.isna(ip_address) or str(ip_address).strip() == '':
            continue
        
        ip_address = str(ip_address).strip()
        hostname = str(hostname).strip() if pd.notna(hostname) else ''
        
        # 检查样式（Excel 行号从1开始，pandas 从0开始）
        excel_row = idx + 1
        excel_col = ip_col_idx + 1
        style_info = style_map.get((excel_row, excel_col), {})
        
        # 过滤删除线
        if exclude_strikethrough and style_info.get('font_strikethrough', False):
            print(f"跳过删除线 IP: {ip_address} ({hostname})")
            continue
        
        # 过滤颜色
        if filter_color == 'green':
            if not is_green_cell(style_info):
                continue
        
        results.append({
            'ip': ip_address,
            'hostname': hostname,
            'row': excel_row
        })
    
    print(f"成功读取 {len(results)} 个 IP 地址")
    return results


def list_available_colors(file_path: str, sheet_name: str = 'net&sec', column_name: str = 'MGMT') -> List[str]:
    """
    列出 sheet 中某一列使用的所有颜色
    
    Args:
        file_path: Excel 文件路径
        sheet_name: sheet 页名称
        column_name: 列名
        
    Returns:
        list: 颜色列表
    """
    # 读取数据找到列索引
    df = read_excel_data_only(file_path, sheet_name)
    
    # 查找列索引
    col_idx = None
    for idx, row in df.iterrows():
        row_values = [str(x).strip() if pd.notna(x) else '' for x in row.values]
        if column_name in row_values:
            col_idx = row_values.index(column_name) + 1  # Excel 列从1开始
            break
    
    if col_idx is None:
        return []
    
    # 读取样式
    style_map = get_sheet_style_map(file_path, sheet_name, max_rows=len(df))
    
    # 收集所有颜色
    colors = set()
    for (row, col), style in style_map.items():
        if col == col_idx and style.get('fill_color'):
            colors.add(style['fill_color'])
    
    return sorted(list(colors))


def find_server_credentials(file_path: str,
                            sheet_name: str = 'server&security',
                            identifier: str = None) -> Optional[Dict]:
    """
    根据 hostname 或管理网IP地址查找服务器的登录凭据
    
    Args:
        file_path: Excel 文件路径
        sheet_name: Sheet 页名称（默认 'server&security'）
        identifier: hostname 或 IP 地址（用于定位服务器行）
        
    Returns:
        dict: 服务器信息
            {
                'hostname': '服务器名称',
                'mgmt_ip': '管理网地址',
                'username': 'System User',
                'password': 'System Password',
                'row': 行号
            }
        如果未找到返回 None
    """
    if not identifier:
        return None
    
    print(f"正在读取 {file_path} 的 {sheet_name} sheet...")
    df = read_excel_data_only(file_path, sheet_name)
    
    # 查找表头行
    header_row_idx = None
    hostname_col_idx = None
    mgmt_ip_col_idx = None
    username_col_idx = None
    password_col_idx = None
    
    for idx, row in df.iterrows():
        row_values = [str(x).strip() if pd.notna(x) else '' for x in row.values]
        
        # 查找所需的列
        if 'hostname' in row_values:
            header_row_idx = idx
            hostname_col_idx = row_values.index('hostname')
            
            # 查找管理网地址列（可能是 '管理网地址' 或 'MGMT'）
            if '管理网地址' in row_values:
                mgmt_ip_col_idx = row_values.index('管理网地址')
            elif 'MGMT' in row_values:
                mgmt_ip_col_idx = row_values.index('MGMT')
            
            # 查找用户名和密码列
            if 'System User' in row_values:
                username_col_idx = row_values.index('System User')
            if 'System Password' in row_values:
                password_col_idx = row_values.index('System Password')
            
            print(f"找到表头行: 第{header_row_idx}行")
            print(f"  hostname 列索引: {hostname_col_idx}")
            print(f"  管理网IP 列索引: {mgmt_ip_col_idx}")
            print(f"  System User 列索引: {username_col_idx}")
            print(f"  System Password 列索引: {password_col_idx}")
            break
    
    if header_row_idx is None:
        raise ValueError(f"在 {sheet_name} sheet 中未找到包含 'hostname' 的表头行")
    
    if username_col_idx is None or password_col_idx is None:
        raise ValueError(f"在 {sheet_name} sheet 中未找到 'System User' 或 'System Password' 列")
    
    # 在数据行中查找匹配的服务器
    data_start_row = header_row_idx + 1
    identifier_lower = identifier.strip().lower()
    
    print(f"开始查找服务器: {identifier}")
    for idx in range(data_start_row, len(df)):
        row = df.iloc[idx]
        hostname = row.iloc[hostname_col_idx]
        mgmt_ip = row.iloc[mgmt_ip_col_idx] if mgmt_ip_col_idx is not None else None
        
        # 跳过空行
        if pd.isna(hostname) and (mgmt_ip is None or pd.isna(mgmt_ip)):
            continue
        
        hostname_str = str(hostname).strip() if pd.notna(hostname) else ''
        mgmt_ip_str = str(mgmt_ip).strip() if mgmt_ip is not None and pd.notna(mgmt_ip) else ''
        
        # 匹配 hostname 或管理网IP
        if (hostname_str.lower() == identifier_lower or 
            mgmt_ip_str.lower() == identifier_lower):
            
            # 读取用户名和密码
            username = row.iloc[username_col_idx]
            password = row.iloc[password_col_idx]
            
            username_str = str(username).strip() if pd.notna(username) else ''
            password_str = str(password).strip() if pd.notna(password) else ''
            
            if not username_str:
                print(f"✗ 找到服务器 {hostname_str}，但 System User 为空")
                continue
            
            if not password_str:
                print(f"✗ 找到服务器 {hostname_str}，但 System Password 为空")
                continue
            
            excel_row = idx + 1
            print(f"✓ 找到服务器: {hostname_str} ({mgmt_ip_str})")
            
            return {
                'hostname': hostname_str,
                'mgmt_ip': mgmt_ip_str,
                'username': username_str,
                'password': password_str,
                'row': excel_row
            }
    
    print(f"✗ 未找到匹配的服务器: {identifier}")
    return None
